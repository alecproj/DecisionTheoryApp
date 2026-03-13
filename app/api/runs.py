from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv
import os

from bson import ObjectId
from flask import Blueprint, request
from openpyxl import load_workbook

from app.algorithms.registry import get_algorithm
from app.reporting.reporter import MarkdownReporter
from app.db.mongo import inputs_col, reports_col

bp = Blueprint("runs_api", __name__, url_prefix="/api")


def xlsx_bytes_to_csv_text(file_bytes: bytes) -> str:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    output = StringIO()
    writer = csv.writer(output, lineterminator="\n")

    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if value is None else str(value) for value in row])

    return output.getvalue()


@bp.post("/runs/<algorithm_id>")
def runs_create(algorithm_id: str):
    report_name = request.form.get("report_name", "").strip()
    if not report_name:
        return {"error": "Не передано имя отчета", "code": "REPORT_NAME_MISSING"}, 400

    if "file" not in request.files:
        return {"error": "Файл не передан", "code": "FILE_MISSING"}, 400

    file = request.files["file"]

    if file.filename == "":
        return {"error": "Файл не выбран", "code": "FILE_MISSING"}, 400

    filename = file.filename
    ext = os.path.splitext(filename)[1].lower()

    if ext not in {".csv", ".xlsx"}:
        return {
            "error": "Неверный формат файла. Ожидается CSV или XLSX",
            "code": "INVALID_FILE_FORMAT"
        }, 400

    file_bytes = file.read()

    if not file_bytes:
        return {"error": "Файл пустой", "code": "FILE_EMPTY"}, 400

    try:
        algo = get_algorithm(algorithm_id)
    except KeyError:
        return {"error": f"Алгоритм '{algorithm_id}' не найден", "code": "ALGORITHM_NOT_FOUND"}, 404

    try:
        if ext == ".csv":
            file_content = file_bytes.decode("utf-8-sig")
        elif ext == ".xlsx":
            file_content = xlsx_bytes_to_csv_text(file_bytes)
        else:
            return {
                "error": "Неподдерживаемый формат файла",
                "code": "INVALID_FILE_FORMAT"
            }, 400
    except UnicodeDecodeError:
        return {
            "error": "CSV должен быть в кодировке UTF-8",
            "code": "INVALID_ENCODING"
        }, 400
    except Exception as e:
        return {
            "error": f"Ошибка чтения Excel файла: {e}",
            "code": "EXCEL_READ_ERROR"
        }, 400

    try:
        typed_input = algo.validate(file_content)
    except ValueError as e:
        return {"error": str(e), "code": "VALIDATION_ERROR"}, 400

    try:
        reporter = MarkdownReporter()
        reporter.h1(report_name)
        algo.run(typed_input, reporter)
        md = reporter.get_markdown()
    except Exception as e:
        return {"error": str(e), "code": "DOMAIN_VALIDATION_ERROR"}, 422

    now = datetime.now(timezone.utc)
    run_id = ObjectId()

    inputs_col().insert_one({
        "algorithm_id": algo.id,
        "run_id": run_id,
        "filename": filename,
        "file": file_content,
        "created_at": now,
    })

    reports_col().insert_one({
        "algorithm_id": algo.id,
        "run_id": run_id,
        "report_name": report_name,
        "markdown": md,
        "created_at": now,
    })

    return {
        "algorithm_id": algorithm_id,
        "run_id": str(run_id),
        "report_name": report_name,
    }, 201

